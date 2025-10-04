"""
Export utilities for test results
"""
import csv
import json
from datetime import datetime
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def export_to_csv(results: List[Dict], filename: str = None) -> str:
    """Export results to CSV file"""
    if not filename:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'test_results_{timestamp}.csv'
    
    with open(filename, 'w', newline='') as csvfile:
        if not results:
            return filename
        
        fieldnames = ['phone_number', 'status', 'timestamp', 'duration', 'error']
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        
        writer.writeheader()
        for result in results:
            writer.writerow({
                'phone_number': result.get('phone_number', ''),
                'status': result.get('status', ''),
                'timestamp': result.get('timestamp', ''),
                'duration': f"{result.get('duration', 0):.2f}s",
                'error': result.get('error', '')
            })
    
    return filename


def export_to_excel(results: List[Dict], filename: str = None) -> str:
    """Export results to Excel file with formatting"""
    if not filename:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'test_results_{timestamp}.xlsx'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Results"
    
    # Add headers with styling
    headers = ['Phone Number', 'Status', 'Timestamp', 'Duration (s)', 'Error']
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add data with conditional formatting
    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    for row, result in enumerate(results, 2):
        ws.cell(row=row, column=1, value=result.get('phone_number', ''))
        
        status_cell = ws.cell(row=row, column=2, value=result.get('status', ''))
        if result.get('status') == 'connected':
            status_cell.fill = success_fill
        else:
            status_cell.fill = fail_fill
        
        ws.cell(row=row, column=3, value=result.get('timestamp', ''))
        ws.cell(row=row, column=4, value=round(result.get('duration', 0), 2))
        ws.cell(row=row, column=5, value=result.get('error', ''))
    
    # Add summary sheet
    summary_ws = wb.create_sheet("Summary")
    summary_ws.cell(row=1, column=1, value="Test Summary").font = Font(bold=True, size=14)
    
    total_tested = len(results)
    connected = len([r for r in results if r.get('status') == 'connected'])
    failed = total_tested - connected
    
    summary_data = [
        ["Total Numbers Tested:", total_tested],
        ["Connected:", connected],
        ["Failed:", failed],
        ["Success Rate:", f"{(connected/total_tested*100 if total_tested > 0 else 0):.1f}%"]
    ]
    
    for row, (label, value) in enumerate(summary_data, 3):
        summary_ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        summary_ws.cell(row=row, column=2, value=value)
    
    # Add connected numbers list
    summary_ws.cell(row=8, column=1, value="Connected Numbers:").font = Font(bold=True)
    connected_numbers = [r.get('phone_number') for r in results if r.get('status') == 'connected']
    
    for idx, number in enumerate(connected_numbers, 9):
        summary_ws.cell(row=idx, column=1, value=number)
    
    # Auto-adjust column widths
    for sheet in wb.worksheets:
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width
    
    wb.save(filename)
    return filename


def export_connected_only(results: List[Dict], format: str = 'csv') -> str:
    """Export only connected phone numbers"""
    connected_results = [r for r in results if r.get('status') == 'connected']
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    if format == 'xlsx':
        filename = f'connected_numbers_{timestamp}.xlsx'
        wb = Workbook()
        ws = wb.active
        ws.title = "Connected Numbers"
        
        # Headers
        ws.cell(row=1, column=1, value="Phone Number").font = Font(bold=True)
        ws.cell(row=1, column=2, value="Test Time").font = Font(bold=True)
        
        # Data
        for row, result in enumerate(connected_results, 2):
            ws.cell(row=row, column=1, value=result.get('phone_number', ''))
            ws.cell(row=row, column=2, value=result.get('timestamp', ''))
        
        wb.save(filename)
    else:
        filename = f'connected_numbers_{timestamp}.csv'
        with open(filename, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(['Phone Number', 'Test Time'])
            for result in connected_results:
                writer.writerow([result.get('phone_number', ''), result.get('timestamp', '')])
    
    return filename